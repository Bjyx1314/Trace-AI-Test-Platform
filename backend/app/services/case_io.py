"""测试用例导入/导出：导出 Markdown / Excel；导入 xmind / Excel / Markdown / Word(docx)。

统一的"用例 dict"形状(供导入产出 / 导出消费)：
  {title, priority, case_type, modules:[..], platforms:[..], preconditions:[..],
   steps:[{seq, action, expected}], expected_result}
导入解析尽量宽松，缺字段给默认值；解析不出步骤时至少保留标题。
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from typing import Any

_DEF_PRIORITY = "P2"
_DEF_TYPE = "功能"


# ─────────────────────────── 导出 ───────────────────────────

def cases_to_markdown(cases: list[Any]) -> str:
    out: list[str] = ["# 测试用例", ""]
    for c in cases:
        cid = getattr(c, "case_id", "") or ""
        out.append(f"## {cid} {getattr(c, 'title', '')}".strip())
        out.append(f"- 优先级: {getattr(c, 'priority', '') or ''}")
        out.append(f"- 模块: {', '.join(getattr(c, 'modules', None) or [])}")
        out.append(f"- 端: {', '.join(getattr(c, 'platforms', None) or [])}")
        out.append(f"- 场景类型: {getattr(c, 'case_type', '') or ''}")
        out.append(f"- 前置条件: {'; '.join(getattr(c, 'preconditions', None) or [])}")
        out.append(f"- 预期结果: {getattr(c, 'expected_result', '') or ''}")
        out.append("")
        steps = getattr(c, "steps", None) or []
        if steps:
            out.append("| # | 操作 | 预期 |")
            out.append("| --- | --- | --- |")
            for i, s in enumerate(steps, 1):
                act = str(s.get("action", "")).replace("\n", " ").replace("|", "/")
                exp = str(s.get("expected", "")).replace("\n", " ").replace("|", "/")
                out.append(f"| {s.get('seq', i)} | {act} | {exp} |")
            out.append("")
    return "\n".join(out)


def cases_to_xlsx(cases: list[Any]) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    headers = ["用例编号", "标题", "优先级", "模块", "端", "场景类型", "前置条件", "步骤", "预期结果"]
    ws.append(headers)
    for c in cases:
        steps = getattr(c, "steps", None) or []
        steps_txt = "\n".join(
            f"{s.get('seq', i)}. {s.get('action', '')} -> {s.get('expected', '')}"
            for i, s in enumerate(steps, 1)
        )
        ws.append([
            getattr(c, "case_id", "") or "",
            getattr(c, "title", "") or "",
            getattr(c, "priority", "") or "",
            ", ".join(getattr(c, "modules", None) or []),
            ", ".join(getattr(c, "platforms", None) or []),
            getattr(c, "case_type", "") or "",
            "; ".join(getattr(c, "preconditions", None) or []),
            steps_txt,
            getattr(c, "expected_result", "") or "",
        ])
    # 列宽
    for col, w in zip("ABCDEFGHI", (14, 36, 8, 16, 16, 10, 24, 50, 30)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── 导入 ───────────────────────────

def parse_import(filename: str, data: bytes) -> list[dict]:
    """按扩展名分派解析，返回用例 dict 列表。"""
    name = (filename or "").lower()
    if name.endswith(".xmind"):
        return _parse_xmind(data)
    if name.endswith((".xlsx", ".xls")):
        return _parse_xlsx(data)
    if name.endswith((".md", ".markdown")):
        return _parse_markdown(data.decode("utf-8", "ignore"))
    if name.endswith((".docx", ".doc")):
        return _parse_docx(data)
    raise ValueError("不支持的文件格式，请用 xmind / xlsx / md / docx")


def _norm(case: dict) -> dict:
    """补默认值，规整 steps。"""
    steps = case.get("steps") or []
    norm_steps = []
    for i, s in enumerate(steps, 1):
        if isinstance(s, dict):
            norm_steps.append({"seq": i, "action": str(s.get("action", "")), "expected": str(s.get("expected", ""))})
        else:
            norm_steps.append({"seq": i, "action": str(s), "expected": ""})
    return {
        "title": (case.get("title") or "").strip() or "未命名用例",
        "priority": case.get("priority") or _DEF_PRIORITY,
        "case_type": case.get("case_type") or _DEF_TYPE,
        "modules": case.get("modules") or [],
        "platforms": case.get("platforms") or [],
        "preconditions": case.get("preconditions") or [],
        "steps": norm_steps,
        "expected_result": case.get("expected_result") or "",
    }


def _split(text: str) -> list[str]:
    return [x.strip() for x in re.split(r"[,，;；/]", text or "") if x.strip()]


def _parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    ci = {
        "title": col("标题", "用例标题", "title", "name"),
        "priority": col("优先级", "priority"),
        "modules": col("模块", "modules", "module"),
        "platforms": col("端", "适用端", "platforms", "platform"),
        "case_type": col("场景类型", "用例类型", "case_type", "type"),
        "pre": col("前置条件", "preconditions"),
        "steps": col("步骤", "测试步骤", "steps"),
        "expected": col("预期结果", "预期", "expected_result", "expected"),
    }
    out: list[dict] = []
    for r in rows[1:]:
        if not r or ci["title"] < 0 or not (r[ci["title"]] if ci["title"] < len(r) else None):
            continue
        g = lambda k: (str(r[ci[k]]) if 0 <= ci[k] < len(r) and r[ci[k]] is not None else "")
        steps = _parse_steps_text(g("steps"))
        out.append(_norm({
            "title": g("title"), "priority": g("priority").strip() or None,
            "case_type": g("case_type").strip() or None,
            "modules": _split(g("modules")), "platforms": _split(g("platforms")),
            "preconditions": _split(g("pre")), "steps": steps,
            "expected_result": g("expected"),
        }))
    return out


def _parse_steps_text(text: str) -> list[dict]:
    """解析步骤单元格：每行 '1. 操作 -> 预期'(-> 可缺)。"""
    steps = []
    for line in (text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"^\d+[.、)]\s*", "", line)
        if "->" in line:
            act, exp = line.split("->", 1)
        elif "=>" in line:
            act, exp = line.split("=>", 1)
        else:
            act, exp = line, ""
        steps.append({"action": act.strip(), "expected": exp.strip()})
    return steps


def _parse_markdown(text: str) -> list[dict]:
    """解析 cases_to_markdown 导出的格式：## 标题 + 字段bullet + 步骤表。"""
    out: list[dict] = []
    blocks = re.split(r"\n(?=##\s)", text)
    for b in blocks:
        m = re.match(r"##\s+(.*)", b.strip())
        if not m:
            continue
        title = m.group(1).strip()
        title = re.sub(r"^(TC-[A-Za-z0-9-]+)\s+", "", title)  # 去掉前导编号
        cur = {"title": title, "steps": []}
        for line in b.splitlines():
            ls = line.strip()
            if ls.startswith("- 优先级:"):
                cur["priority"] = ls.split(":", 1)[1].strip() or None
            elif ls.startswith("- 模块:"):
                cur["modules"] = _split(ls.split(":", 1)[1])
            elif ls.startswith("- 端:"):
                cur["platforms"] = _split(ls.split(":", 1)[1])
            elif ls.startswith("- 场景类型:"):
                cur["case_type"] = ls.split(":", 1)[1].strip() or None
            elif ls.startswith("- 前置条件:"):
                cur["preconditions"] = _split(ls.split(":", 1)[1])
            elif ls.startswith("- 预期结果:"):
                cur["expected_result"] = ls.split(":", 1)[1].strip()
            elif ls.startswith("|") and "---" not in ls:
                cells = [c.strip() for c in ls.strip("|").split("|")]
                if len(cells) >= 3 and cells[0] not in ("#", "序号") and cells[1] != "操作":
                    cur["steps"].append({"action": cells[1], "expected": cells[2]})
        out.append(_norm(cur))
    return out


def _parse_docx(data: bytes) -> list[dict]:
    """Word：标题样式段落(Heading)作为用例标题，其余段落作为步骤(操作)。表格行也尝试解析为步骤。"""
    import docx
    doc = docx.Document(io.BytesIO(data))
    out: list[dict] = []
    cur: dict | None = None
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if not t:
            continue
        is_heading = (p.style and p.style.name and "Heading" in p.style.name) or bool(re.match(r"^(用例|测试用例|TC[-：:])", t))
        if is_heading:
            if cur:
                out.append(_norm(cur))
            cur = {"title": re.sub(r"^(用例|测试用例)[:：]?\s*", "", t), "steps": []}
        elif cur is not None:
            line = re.sub(r"^\d+[.、)]\s*", "", t)
            if "->" in line or "=>" in line:
                sep = "->" if "->" in line else "=>"
                act, exp = line.split(sep, 1)
                cur["steps"].append({"action": act.strip(), "expected": exp.strip()})
            else:
                cur["steps"].append({"action": line, "expected": ""})
    if cur:
        out.append(_norm(cur))
    return [c for c in out if c["title"]]


def _parse_xmind(data: bytes) -> list[dict]:
    """xmind 本质是 zip：优先 content.json(新版/zen)，回退 content.xml(旧版)。
    约定：根主题下的一级子主题=用例标题；其子主题=步骤(操作)，步骤的子主题(若有)=预期。"""
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        names = z.namelist()
        if "content.json" in names:
            return _xmind_from_json(json.loads(z.read("content.json").decode("utf-8", "ignore")))
        if "content.xml" in names:
            return _xmind_from_xml(z.read("content.xml"))
    raise ValueError("无法解析该 xmind 文件(缺 content.json/content.xml)")


def _xmind_topic_children(topic: dict) -> list[dict]:
    ch = (topic.get("children") or {}).get("attached") or []
    return ch


def _xmind_from_json(doc: Any) -> list[dict]:
    sheets = doc if isinstance(doc, list) else [doc]
    out: list[dict] = []
    for sheet in sheets:
        root = sheet.get("rootTopic") or {}
        for case_topic in _xmind_topic_children(root):
            title = (case_topic.get("title") or "").strip()
            if not title:
                continue
            steps = []
            for step_topic in _xmind_topic_children(case_topic):
                act = (step_topic.get("title") or "").strip()
                exp_children = _xmind_topic_children(step_topic)
                exp = (exp_children[0].get("title") or "").strip() if exp_children else ""
                steps.append({"action": act, "expected": exp})
            out.append(_norm({"title": title, "steps": steps}))
    return out


def _xmind_from_xml(raw: bytes) -> list[dict]:
    from lxml import etree
    tree = etree.fromstring(raw)
    ns = {"x": tree.nsmap.get(None, "urn:xmind:xmap:xmlns:content:2.0")}

    def title_of(topic):
        t = topic.find("x:title", ns)
        return (t.text or "").strip() if t is not None else ""

    def children_of(topic):
        ch = topic.find("x:children", ns)
        if ch is None:
            return []
        topics = ch.find("x:topics", ns)
        return topics.findall("x:topic", ns) if topics is not None else []

    out: list[dict] = []
    for sheet in tree.findall("x:sheet", ns):
        root = sheet.find("x:topic", ns)
        if root is None:
            continue
        for case_topic in children_of(root):
            title = title_of(case_topic)
            if not title:
                continue
            steps = []
            for step_topic in children_of(case_topic):
                exp_nodes = children_of(step_topic)
                steps.append({"action": title_of(step_topic),
                              "expected": title_of(exp_nodes[0]) if exp_nodes else ""})
            out.append(_norm({"title": title, "steps": steps}))
    return out
